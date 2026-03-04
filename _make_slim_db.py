"""slim DB テンプレート生成スクリプト"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLS = ['開催日','開催場','レース番号','車番','選手名',
        'IP','EP','DP','BP','直線の伸び','戦法','is_monster','is_unreliable']
NOTES = ['例: 2026/03/04','例: 立川','例: 7','1〜9','例: 田中誠',
         '1〜10','1〜10','1〜10','1〜10','S/A/B/C',
         '逃げ切り/差し等','0 or 1','0 or 1']
WIDTHS = [13,10,12,6,12,6,6,6,6,12,16,12,14]
SAMPLE = ['2026/03/04','立川',7,3,'田中誠',7,6,4,5,'A','先行',0,0]

wb = Workbook()
for i, sheet_name in enumerate(['F1', 'G3~1']):
    ws = wb.active if i == 0 else wb.create_sheet(sheet_name)
    ws.title = sheet_name

    hdr_fill  = PatternFill(fill_type='solid', fgColor='1a1a2e')
    note_fill = PatternFill(fill_type='solid', fgColor='2d2d44')

    for ci, (col, note, w) in enumerate(zip(COLS, NOTES, WIDTHS), 1):
        hc = ws.cell(row=1, column=ci, value=col)
        hc.font = Font(bold=True, color='FFCC00', size=10)
        hc.fill = hdr_fill
        hc.alignment = Alignment(horizontal='center')

        nc = ws.cell(row=2, column=ci, value=note)
        nc.font = Font(italic=True, color='88aacc', size=9)
        nc.fill = note_fill
        nc.alignment = Alignment(horizontal='center')

        sc = ws.cell(row=3, column=ci, value=SAMPLE[ci-1])
        sc.alignment = Alignment(horizontal='center')

        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = 'A3'

out = 'data/S級DB_slim.xlsx'
wb.save(out)
print(f'完了: {out}')
